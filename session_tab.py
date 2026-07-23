"""
session_tab.py — Testability Scanner  v1.2  (i18n)
────────────────────────────────────────────────────────────────
Multi-page session tarama mantığı.
Tüm log ve hata mesajları i18n.t() üzerinden yönetilir.

v1.2 değişiklikleri:
  - Aynı flow adıyla tekrar çalıştırıldığında Excel / Word / JSON
    tutarlı biçimde "oturum N" suffix'i alır (strings.json: session_run_label).
  - JSON dosya adından tarih-saat damgası kaldırıldı.
  - Hardcoded "Unique ID" / "Undefined ID" / "Duplicate ID" / "Missing ID"
    string literal'leri kaldırıldı — artık constants.py'deki STATUS_* ve
    STATUS_ORDER sabitleri kullanılıyor.
"""

import threading
import os
from datetime import datetime
from collections import Counter
from i18n import t
from constants import (
    STATUS_UNIQUE, STATUS_UNDEFINED, STATUS_DUPLICATE, STATUS_MISSING,
    STATUS_ORDER,
)


class SessionTab:

    def __init__(self, app_ref):
        self.app = app_ref

        self._driver         = None
        self._platform       = "ios"
        self._pixel_ratio    = 1.0
        self._running        = False
        self._collecting     = False

        self._flow_name      = ""
        self._scan_counter   = 0

        self._all_elements:  list[dict] = []
        self._page_summary:  list[dict] = []

        self._log_box         = None
        self._update_table_cb = None

    # ── Bağlama ──────────────────────────────────────────────────────────────
    def bind_log(self, log_box):
        self._log_box = log_box

    def bind_table_callback(self, cb):
        self._update_table_cb = cb

    # ── Log ──────────────────────────────────────────────────────────────────
    def _log(self, text, tag=""):
        if not self._log_box:
            return
        def _d():
            self._log_box.configure(state="normal")
            ts = datetime.now().strftime("%H:%M:%S")
            self._log_box._textbox.insert("end", f"[{ts}] {text}\n", tag or "")
            self._log_box._textbox.see("end")
            self._log_box.configure(state="disabled")
        self.app.after(0, _d)

    # ── Oturumu Başlat ────────────────────────────────────────────────────────
    def start_session(self, flow_name: str):
        if self._driver:
            self._log(t("session_log_active_session"), "warn")
            return

        if not flow_name.strip():
            self._log(t("session_log_flow_empty"), "err")
            return

        pf = self.app.v_platform.get()
        if pf == "ios":
            profile = self.app.ios_panel.get_data()
            if not profile.get("bundle_id"):
                self._log(t("session_log_ios_empty"), "err")
                return
        else:
            profile = self.app.and_panel.get_data()
            if not profile.get("app_package"):
                self._log(t("session_log_and_empty"), "err")
                return

        output_dir = self.app.v_out_dir.get().strip()
        if not output_dir:
            self._log(t("session_log_output_empty"), "err")
            return

        self._flow_name    = flow_name.strip()
        self._scan_counter = 0
        self._all_elements.clear()
        self._page_summary.clear()
        if self._update_table_cb:
            self.app.after(0, self._update_table_cb)

        self._platform = pf
        self._running  = True
        self.app.after(0, lambda: self.app._set_session_state("connected"))

        self._log(t("session_log_sep"), "dim")
        self._log(t("session_log_connecting", platform=pf.upper()), "info")
        self._log(t("session_log_flow", flow=self._flow_name), "info")

        threading.Thread(
            target=self._connect_worker,
            args=(pf, profile, self.app.v_appium.get()),
            daemon=True
        ).start()

    def _connect_worker(self, platform, profile, appium_server):
        try:
            from appium import webdriver

            if platform == "ios":
                from appium.options.ios import XCUITestOptions
                options = XCUITestOptions()
                options.platform_name    = "iOS"
                options.device_name      = profile["device_name"]
                options.platform_version = profile["platform_version"]
                options.automation_name  = "XCUITest"
                options.bundle_id        = profile["bundle_id"]
                options.no_reset         = profile.get("no_reset", True)
                options.udid             = profile["udid"]
            else:
                from appium.options.android import UiAutomator2Options
                options = UiAutomator2Options()
                options.platform_name    = "Android"
                options.device_name      = profile["device_name"]
                options.platform_version = profile["platform_version"]
                options.automation_name  = "UiAutomator2"
                options.app_package      = profile["app_package"]
                options.app_activity     = profile["app_activity"]
                options.no_reset         = profile.get("no_reset", True)

            self._driver = webdriver.Remote(appium_server, options=options)

            if platform == "ios":
                try:
                    import tempfile, os as _os
                    tmp = tempfile.mktemp(suffix=".png")
                    self._driver.get_screenshot_as_file(tmp)
                    from PIL import Image as PILImage
                    with PILImage.open(tmp) as img:
                        ss_w, _ = img.size
                    _os.remove(tmp)
                    win = self._driver.get_window_size()
                    pt_w = win.get("width", ss_w)
                    self._pixel_ratio = ss_w / pt_w if pt_w > 0 else 1.0
                    self._log(t("session_log_pixel_ratio", ratio=self._pixel_ratio), "dim")
                except Exception:
                    self._pixel_ratio = 1.0
            else:
                self._pixel_ratio = 1.0

            self._log(t("session_log_connected"), "ok")
            self._log(t("session_log_hint1"), "dim")
            self._log(t("session_log_hint2"), "dim")
            self._log(t("session_log_hint3"), "dim")

        except Exception as ex:
            self._log(t("session_log_conn_error", error=ex), "err")
            self._driver  = None
            self._running = False
            self.app.after(0, lambda: self.app._set_session_state("idle"))

    # ── Sayfayı Topla ─────────────────────────────────────────────────────────
    def collect_page(self):
        if not self._driver:
            self._log(t("session_log_ios_empty"), "err")
            return
        if self._collecting:
            self._log(t("session_log_active_session"), "warn")
            return

        self._scan_counter += 1
        scan_label = f"{self._flow_name} - {t('session_scan_label')} {self._scan_counter}"

        self._collecting = True
        self.app.after(0, lambda: self.app._set_session_state("collecting"))
        self._log(t("session_log_scanning", label=scan_label), "info")

        threading.Thread(
            target=self._collect_worker,
            args=(scan_label,),
            daemon=True
        ).start()

    def _collect_worker(self, scan_label: str):
        try:
            import smart_checker as sc

            output_dir = self.app.v_out_dir.get().strip()
            ss_dir     = os.path.join(output_dir, f"screenshots_{self._platform}")
            os.makedirs(ss_dir, exist_ok=True)
            ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_label = scan_label.replace(" ", "_").replace("-", "").replace("/", "_")
            ss_path    = os.path.join(ss_dir, f"{safe_label}_{ts}.png")
            self._driver.get_screenshot_as_file(ss_path)

            win = self._driver.get_window_size()
            sw, sh = win["width"], win["height"]

            elements = sc._collect_elements(
                self._driver, self._platform,
                scan_label,
                sw, sh, self._pixel_ratio,
                self._log
            )

            for e in elements:
                e["page"]      = scan_label
                e["flow_name"] = self._flow_name

            import shared as shr
            elements = shr.enrich_with_ai(elements, self._platform)

            self._all_elements.extend(elements)

            counts = Counter(e["status"] for e in elements)
            summary = {
                "label":     scan_label,
                "flow":      self._flow_name,
                "unique":    counts[STATUS_UNIQUE],
                "undefined": counts[STATUS_UNDEFINED],
                "duplicate": counts[STATUS_DUPLICATE],
                "missing":   counts[STATUS_MISSING],
                "total":     len(elements),
                "ss_path":   ss_path,
            }
            self._page_summary.append(summary)

            self._log(
                t("session_log_scan_ok",
                  label=scan_label,
                  unique=summary["unique"],
                  undefined=summary["undefined"],
                  duplicate=summary["duplicate"],
                  missing=summary["missing"],
                  total=summary["total"]),
                "ok"
            )

            if self._update_table_cb:
                self.app.after(0, self._update_table_cb)

            self.app.after(0, lambda: self.app._set_session_state("connected"))

        except Exception as ex:
            import traceback
            self._log(t("session_log_collect_error", error=ex), "err")
            self._log(traceback.format_exc(), "err")
            self._scan_counter -= 1
            self.app.after(0, lambda: self.app._set_session_state("connected"))
        finally:
            self._collecting = False

    # ── Bitti - Rapor Oluştur ─────────────────────────────────────────────────
    def finish_session(self):
        if not self._all_elements:
            self._log(t("session_log_no_elements"), "warn")
            return

        self.app.after(0, lambda: self.app._set_session_state("finishing"))
        total_scans = len(self._page_summary)
        self._log(t("session_log_sep"), "dim")
        self._log(
            t("session_log_report_gen",
              total=len(self._all_elements),
              scans=total_scans,
              flow=self._flow_name),
            "info"
        )

        threading.Thread(target=self._finish_worker, daemon=True).start()

    # ── Oturum numarası: Word / Excel / JSON için ortak "ilk boş slot" ────────
    def _resolve_session_num(self, output_dir, flow_safe, plat_suffix,
                              excel_file, run_label) -> int:
        """
        Eskiden sadece Excel sheet sayısına bakılıyordu. Sorun: önceki bir
        oturumda örn. sadece Excel üretilmiş (Word kapalıyken), bir sonraki
        oturumda session_num yine 1 hesaplanıyor ve Word, önceki oturumun
        dosyasının üzerine yazıyordu.

        Artık Word dosyası / JSON dosyası / Excel sheet'i birlikte kontrol
        edilip, bu flow + platform için üçünden de HİÇBİRİNİN var olmadığı
        ilk N bulunur. Böylece hangi formatlar daha önce üretilmiş olursa
        olsun, Word asla eski bir oturumun dosyasının üzerine yazılmaz.
        """
        n = 1
        while True:
            suffix     = "" if n == 1 else f" - {run_label} {n}"
            sheet_name = self._flow_name + suffix
            word_file  = os.path.join(
                output_dir, f"Session_{flow_safe}{suffix}_{plat_suffix}.docx")
            json_file  = os.path.join(
                output_dir, f"Session_{flow_safe}{suffix}_{self._platform}.json")

            word_exists = os.path.exists(word_file)
            json_exists = os.path.exists(json_file)

            sheet_exists = False
            if os.path.exists(excel_file):
                try:
                    import openpyxl as _opxl
                    wb_check = _opxl.load_workbook(excel_file, read_only=True)
                    sheet_exists = sheet_name in wb_check.sheetnames
                    wb_check.close()
                except Exception:
                    sheet_exists = False

            if not (word_exists or json_exists or sheet_exists):
                return n
            n += 1

    def _finish_worker(self):
        try:
            if self._driver:
                try:
                    self._driver.quit()
                    self._log(t("session_log_driver_closed"), "dim")
                except Exception:
                    pass
                self._driver = None

            import shared as shr

            output_dir   = self.app.v_out_dir.get().strip()
            cfg          = self.app._collect()
            fmt_parts    = set()
            if cfg.get("output_word"):  fmt_parts.add("word")
            if cfg.get("output_excel"): fmt_parts.add("excel")
            if cfg.get("output_json"):  fmt_parts.add("json")

            doc_sections = cfg.get("document_sections",
                                   ["unique", "undefined", "duplicate", "missing"])
            plat_suffix  = "IOS" if self._platform == "ios" else "Android"
            flow_safe    = self._flow_name.replace(" ", "_")

            # ── Oturum numarasını belirle ─────────────────────────────────────
            # Sadece Excel sheet sayısına bakmak yeterli değil: bir önceki
            # oturumda örn. sadece Excel üretilmiş, Word üretilmemiş olabilir.
            # Bu yüzden Word / Excel / JSON üçü birlikte kontrol edilip bu
            # flow + platform için "ilk boş slot" bulunur. Böylece Word hiçbir
            # zaman önceki bir oturumun dosyasının üzerine yazılmaz.
            excel_file = os.path.join(
                output_dir, f"Session_{flow_safe}_{plat_suffix}.xlsx")
            run_label   = t("session_run_label")
            session_num = self._resolve_session_num(
                output_dir, flow_safe, plat_suffix, excel_file, run_label)

            # "oturum 1" değil, ilk çalışmada suffix yok — aynı Excel davranışı
            session_suffix = ("" if session_num == 1
                              else f" - {run_label} {session_num}")
            sheet_name = self._flow_name + session_suffix

            self._log(t("session_log_sheet_name", name=sheet_name), "dim")

            # ── Excel ─────────────────────────────────────────────────────────
            if "excel" in fmt_parts:
                first_scan = True
                for s in self._page_summary:
                    scan_elems = [e for e in self._all_elements
                                  if e.get("page") == s["label"]]
                    scan_elems.sort(
                        key=lambda e: STATUS_ORDER.get(e.get("status", ""), 99))

                    if first_scan:
                        shr.generate_excel(
                            all_elements=scan_elems,
                            page_name=sheet_name,
                            excel_file=excel_file,
                            document_sections=doc_sections,
                            platform=self._platform,
                            screenshot_path="",
                        )
                        first_scan = False
                    else:
                        self._append_scan_rows(
                            excel_file, sheet_name, scan_elems,
                            doc_sections, shr)

                self._append_screenshots_to_excel(excel_file, sheet_name)
                self._log(t("session_log_excel_saved",
                             file=excel_file, sheet=sheet_name), "ok")

            # ── Word ──────────────────────────────────────────────────────────
            if "word" in fmt_parts:
                # Tüm scan'ler için tek birleşik Word dosyası (flow bazlı).
                # NOT: shared.generate_word her çağrıda dosyayı sıfırdan
                # oluşturuyor. Eskiden döngü içinde sayfa başına çağrılıyordu
                # ve sonuçta dosyada sadece SON sayfa kalıyordu. Artık tüm
                # oturumun elementleri birleştirilip TEK seferde yazılıyor.
                word_file = os.path.join(
                    output_dir,
                    f"Session_{flow_safe}{session_suffix}_{plat_suffix}.docx")

                combined_elems = []
                for s in self._page_summary:
                    combined_elems.extend(
                        e for e in self._all_elements
                        if e.get("page") == s["label"])

                first_ss = next(
                    (s.get("ss_path", "") for s in self._page_summary
                     if s.get("ss_path")), "")

                shr.generate_word(
                    all_elements=combined_elems,
                    page_name=self._flow_name,
                    word_file=word_file,
                    document_sections=doc_sections,
                    platform=self._platform,
                    screenshot_path=first_ss,
                )
                self._log(t("session_log_word_saved",
                             file=os.path.basename(word_file)), "ok")

            # ── JSON ──────────────────────────────────────────────────────────
            if "json" in fmt_parts:
                json_file = os.path.join(
                    output_dir,
                    f"Session_{flow_safe}{session_suffix}_{self._platform}.json")
                shr.generate_json(
                    elements=self._all_elements,
                    page_name=self._flow_name,
                    json_file=json_file,
                    platform=self._platform,
                )
                self._log(t("session_log_json_saved",
                             file=os.path.basename(json_file)), "ok")

            self._log(t("session_log_sep"), "dim")
            self._log(
                t("session_log_done",
                  flow=self._flow_name,
                  scans=len(self._page_summary),
                  total=len(self._all_elements)),
                "ok"
            )
            self.app.after(0, lambda: self.app.badge.set("ok"))

        except Exception as ex:
            import traceback
            self._log(t("session_log_report_error", error=ex), "err")
            self._log(traceback.format_exc(), "err")
            self.app.after(0, lambda: self.app.badge.set("error"))
        finally:
            self._running    = False
            self._collecting = False
            self.app.after(0, lambda: self.app._set_session_state("idle"))

    # ── Sonraki tarama satırlarını mevcut sheet'e ekle ───────────────────────
    def _append_scan_rows(self, excel_file, sheet_name, scan_elems,
                           doc_sections, shr):
        import openpyxl
        from openpyxl.styles import Font, Alignment

        wb  = openpyxl.load_workbook(excel_file)
        ws  = wb[sheet_name]

        last_row = ws.max_row + 1
        ordered  = shr._build_ordered(scan_elems, doc_sections)

        data_rows_so_far = max(0, last_row - 3)

        COL_KEYS = ["element_id", "page", "type", "label", "value",
                    "acc_id", "status", "new_status", "ai_suggestion"]

        for idx, elem in enumerate(ordered):
            elem_id    = f"{sheet_name}_element_{data_rows_so_far + idx + 1}"
            status     = elem.get("status", STATUS_MISSING)
            new_status = shr.get_new_status(status)
            row_num    = last_row + idx
            pal        = shr.STATUS_PALETTE.get(status, shr.STATUS_PALETTE[STATUS_MISSING])
            r_fill     = shr.fill(pal["row"] if idx % 2 == 0 else pal["alt"])
            ns_fill    = shr.fill(shr.NEW_STATUS_COLOR["row"] if idx % 2 == 0
                                  else shr.NEW_STATUS_COLOR["alt"])
            ai_fill    = shr.fill(shr.AI_SUGGESTION_COLOR["row"] if idx % 2 == 0
                                  else shr.AI_SUGGESTION_COLOR["alt"])

            for ci, key in enumerate(COL_KEYS, 1):
                val = shr._get_val(elem, key, elem_id, new_status)
                c   = ws.cell(row=row_num, column=ci, value=val)
                c.border = shr.BORDER
                shr._style_excel_cell(c, key, pal, r_fill, ns_fill,
                                       ai_fill, new_status)

            ws.row_dimensions[row_num].height = 60

        shr.safe_save(wb, excel_file)

    # ── Screenshotları Excel'e ekle ───────────────────────────────────────────
    def _append_screenshots_to_excel(self, excel_file: str, sheet_name: str):
        try:
            import openpyxl
            from openpyxl.drawing.image import Image as XLImage
            from PIL import Image as PILImage
            import shared as shr

            wb = openpyxl.load_workbook(excel_file)
            if sheet_name not in wb.sheetnames:
                self._log(t("session_log_sheet_missing", sheet=sheet_name), "warn")
                return
            ws = wb[sheet_name]

            data_cols   = ws.max_column
            img_col_idx = data_cols + 2
            from openpyxl.utils import get_column_letter
            img_col_ltr = get_column_letter(img_col_idx)

            ws.column_dimensions[get_column_letter(data_cols + 1)].width = 2
            ws.column_dimensions[img_col_ltr].width = 44

            current_row = 1
            tmp_files   = []

            for s in self._page_summary:
                ss_path = s.get("ss_path", "")
                label   = s.get("label", "tarama")

                if not ss_path or not os.path.exists(ss_path):
                    continue

                try:
                    with PILImage.open(ss_path) as img:
                        orig_w, orig_h = img.size
                    target_w = 300
                    target_h = int(orig_h * target_w / orig_w)

                    tmp_path = ss_path.replace(".png", "_xl_tmp.png")
                    with PILImage.open(ss_path) as img:
                        img.resize((target_w, target_h),
                                   PILImage.LANCZOS).save(tmp_path, "PNG")
                    tmp_files.append(tmp_path)

                    label_cell = ws.cell(row=current_row, column=img_col_idx,
                                         value=f"📸 {label}")
                    label_cell.font      = openpyxl.styles.Font(
                        bold=True, color="1F4E79", size=10)
                    label_cell.fill      = openpyxl.styles.PatternFill(
                        "solid", fgColor="DEEAF1")
                    label_cell.alignment = openpyxl.styles.Alignment(
                        horizontal="center", vertical="center")
                    ws.row_dimensions[current_row].height = 20
                    current_row += 1

                    xl_img        = XLImage(tmp_path)
                    xl_img.width  = target_w
                    xl_img.height = target_h
                    ws.add_image(xl_img, f"{img_col_ltr}{current_row}")

                    rows_needed = max(1, target_h // 15 + 1)
                    for r in range(current_row, current_row + rows_needed):
                        ws.row_dimensions[r].height = 15
                    current_row += rows_needed

                    ws.row_dimensions[current_row].height = 8
                    ws.row_dimensions[current_row + 1].height = 8
                    current_row += 2

                except Exception as e:
                    self._log(t("session_log_screenshot_err",
                                label=label, error=e), "warn")

            shr.safe_save(wb, excel_file)

            for tmp in tmp_files:
                try:
                    os.remove(tmp)
                except OSError:
                    pass

            self._log(t("session_log_screenshots", n=len(self._page_summary)), "dim")

        except Exception as ex:
            self._log(t("session_log_ss_add_error", error=ex), "warn")

    # ── Verileri sıfırla (sadece oturum kapalıyken çağrılmalı) ─────────────────
    def reset(self):
        self._scan_counter = 0
        self._all_elements.clear()
        self._page_summary.clear()
        if self._update_table_cb:
            self.app.after(0, self._update_table_cb)

    # ── Oturumu iptal et ──────────────────────────────────────────────────────
    def abort_session(self):
        if self._driver:
            try:
                self._driver.quit()
            except Exception:
                pass
            self._driver = None
        self._running    = False
        self._collecting = False
        self._log(t("session_log_abort"), "warn")
        self.app.after(0, lambda: self.app._set_session_state("idle"))

    # ── Yardımcılar ───────────────────────────────────────────────────────────
    @property
    def is_active(self):
        return self._driver is not None

    @property
    def scan_count(self):
        return self._scan_counter

    @property
    def flow_name(self):
        return self._flow_name

    @property
    def summary(self):
        return list(self._page_summary)